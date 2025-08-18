#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pre-filler for RFP Excel files.
- Loads an RFP Excel file
- Uses the vector database (Qdrant) to find relevant answers for each question
- Pre-fills the Excel with suggested answers and LLM-generated comments
"""

import openpyxl
from qdrant.client import get_qdrant_client, INTERNAL_COLLECTION
from qdrant_client.http import models as qmodels
from sentence_transformers import SentenceTransformer
import settings
import openai


def get_questions_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    questions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        question = row[0]
        if question:
            questions.append(question)
    return questions, wb, ws


def search_answers_and_contexts(question, embedding_model=None, top_k=3):
    """
    Simplified function for UI testing - returns empty results
    In a real implementation, this would search the vector database
    """
    # For UI testing, return empty/mock results
    answers = [""] * top_k
    contexts = [""] * top_k
    return answers, contexts


def process_rfp_with_ai(dataframe):
    """
    Blank function for UI testing - does nothing but returns the same dataframe
    In a real implementation, this would process the entire RFP with AI
    """
    # For UI testing, just return the same dataframe unchanged
    print("🧪 Mock pre-completion function called (does nothing)")
    return dataframe


def search_answers_and_contexts_full(questions, top_k=3):
    """
    Original function renamed for batch processing
    """
    client = get_qdrant_client()
    model = SentenceTransformer(settings.EMBEDDING_MODEL)
    answers = []
    contexts = []
    for q in questions:
        vec = model.encode([q])[0]
        result = client.search(
            collection_name=INTERNAL_COLLECTION,
            query_vector=vec,
            limit=top_k,
            with_payload=True
        )
        if result:
            best = result[0].payload.get('text', '')
            context = '\n---\n'.join([r.payload.get('text', '') for r in result if r.payload.get('text', '')])
        else:
            best = ''
            context = ''
        answers.append(best)
        contexts.append(context)
    return answers, contexts


def fill_excel_with_answers_and_comments(wb, ws, answers, comments, output_path):
    for i, (answer, comment) in enumerate(zip(answers, comments), start=2):
        ws.cell(row=i, column=2, value=answer)
        ws.cell(row=i, column=3, value=comment)
    wb.save(output_path)


def main():
    import sys
    if len(sys.argv) < 2:
        print("Usage: python pre-filler.py <rfp_excel_file.xlsx>")
        return
    input_path = sys.argv[1]
    output_path = input_path.replace('.xlsx', '_pre_filled.xlsx')
    print(f"🔍 Reading questions from: {input_path}")
    questions, wb, ws = get_questions_from_excel(input_path)
    print(f"✅ {len(questions)} questions extracted")
    print(f"🔎 Searching for answers in the vector database...")
    answers, contexts = search_answers_and_contexts_full(questions)

    # Generate answers and comments with OpenAI LLM
    openai_api_key = getattr(settings, 'OPENAI_API_KEY', None)
    if not openai_api_key:
        print("⚠️ No OpenAI key found in .env, comments and answers will be empty.")
        comments = ['' for _ in questions]
        answers_final = ['' for _ in questions]
    else:
        openai.api_key = openai_api_key
        comments = []
        answers_final = []
        for q, context in zip(questions, contexts):
            # Generate the answer (Yes/No)
            answer_prompt = (
                f"Here is an RFP question:\n\"\"\"{q}\"\"\"\n\n"
                f"Here are internal document excerpts that may help answer:\n{context}\n\n"
                "Based only on the provided context, answer strictly 'Yes' or 'No' to the question. If the context is insufficient, answer 'No'. Reply with only 'Yes' or 'No'."
            )
            try:
                answer_response = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are an expert assistant for RFP responses. Reply only 'Yes' or 'No'."},
                        {"role": "user", "content": answer_prompt}
                    ],
                    temperature=0.0,
                    max_tokens=3
                )
                answer = answer_response.choices[0].message['content'].strip().split()[0].capitalize()
                if answer not in ['Yes', 'No']:
                    answer = 'No'
            except Exception as e:
                answer = f"[LLM error: {e}]"
            answers_final.append(answer)

            # Generate the comment
            comment_prompt = (
                f"Here is an RFP question:\n\"\"\"{q}\"\"\"\n\n"
                f"Here are internal document excerpts that may help answer:\n{context}\n\n"
                "Write a summary or advisory comment for this question, using the provided context if relevant. Be factual, concise, and professional."
            )
            try:
                comment_response = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are an expert assistant for writing RFP responses."},
                        {"role": "user", "content": comment_prompt}
                    ],
                    temperature=0.2,
                    max_tokens=200
                )
                comment = comment_response.choices[0].message['content'].strip() if hasattr(comment_response.choices[0].message, 'content') else comment_response.choices[0].message.get('content', '').strip()
            except Exception as e:
                comment = f"[LLM error: {e}]"
            comments.append(comment)

    print(f"✅ Answers and comments generated, filling the file...")
    fill_excel_with_answers_and_comments(wb, ws, answers_final, comments, output_path)
    print(f"🎉 Pre-filled file saved: {output_path}")


if __name__ == "__main__":
    main()
